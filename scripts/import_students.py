#!/usr/bin/env python3
"""
Importa clientes ativos de uma planilha XLSX da STRONIX para o painel via /admin/api/students/bulk.

Uso:
  python3 scripts/import_students.py <caminho.xlsx> <base_url>
Exemplo:
  python3 scripts/import_students.py ~/Downloads/clientes-01_05_2026.xlsx https://stronix-sdr-production.up.railway.app

Filtra por Situação do contrato == 'Ativo' AND Situação do cliente == 'Ativo'.
Converte telefone (51) 9 9984-9349 em 5551999849349.
"""
import sys
import re
import json
import urllib.request
import openpyxl

def normalize_phone(raw):
    digits = re.sub(r'\D', '', str(raw or ''))
    if len(digits) == 11:
        return '55' + digits
    if len(digits) == 10:
        return '55' + digits[:2] + '9' + digits[2:]
    return None

def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    xlsx_path = sys.argv[1]
    base_url = sys.argv[2].rstrip('/')

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Agrupa por phone — múltiplos clientes ativos no mesmo número (família) viram 1 entrada
    grouped = {}
    skipped = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        nome, _email, contrato, telefone, sit_contrato, sit_cliente = row[0], row[1], row[2], row[3], row[4], row[5]
        if sit_contrato != 'Ativo' or sit_cliente != 'Ativo':
            continue
        phone = normalize_phone(telefone)
        if not phone:
            skipped.append((nome, telefone))
            continue
        if phone not in grouped:
            grouped[phone] = {'names': [], 'contratos': []}
        if nome:
            grouped[phone]['names'].append(str(nome).strip())
        if contrato:
            grouped[phone]['contratos'].append(str(contrato).strip())

    items = []
    for phone, data in grouped.items():
        items.append({
            'phone': phone,
            'name': ' / '.join(data['names']) if data['names'] else None,
            'notes': ' + '.join(set(data['contratos'])) if data['contratos'] else None,
        })

    duplicados = sum(1 for d in grouped.values() if len(d['names']) > 1)
    print(f'Parseados: {len(items)} phones únicos ({duplicados} com múltiplos clientes) | skipped (sem phone): {len(skipped)}')

    if not items:
        print('Nada a importar.')
        sys.exit(0)

    print(f'Amostra (primeiros 3):')
    for it in items[:3]:
        print(f"  {it['phone']} | {it['name']} | {it['notes']}")

    print(f'Enviando POST {base_url}/admin/api/students/bulk ...')
    req = urllib.request.Request(
        f'{base_url}/admin/api/students/bulk',
        data=json.dumps({'items': items}).encode('utf-8'),
        headers={'Content-Type': 'application/json'},
        method='POST',
    )
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            body = resp.read().decode('utf-8')
            print(f'HTTP {resp.status}: {body}')
    except urllib.error.HTTPError as e:
        print(f'HTTP {e.code}: {e.read().decode("utf-8", errors="replace")}')
        sys.exit(1)

if __name__ == '__main__':
    main()
